#!/usr/bin/env python3
"""
Gate A/B/C workbook builder for the Group Dynamics Simulator.

Creates an Excel workbook (.xlsx) with Gate A, Gate B, and Gate C tabs, pre-filled
with the 5-person Alpha Leadership Team synthetic dataset defined in PLAN.md.

Run:   python3 scripts/build_workbook.py
Output: workbook/group-dynamics-simulator-phase1.xlsx
"""

import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ── Version & date ────────────────────────────────────────────────────────────

VERSION = "phase1_gateC_v1"
TODAY   = datetime.date.today().isoformat()

# ── Synthetic dataset ─────────────────────────────────────────────────────────

PEOPLE = [
    {"id": "alex.rivera", "name": "Alex Rivera",  "role": "leader",  "group": "Alpha Leadership Team", "authority": 5, "active": "TRUE"},
    {"id": "jordan.chen", "name": "Jordan Chen",  "role": "manager", "group": "Alpha Leadership Team", "authority": 4, "active": "TRUE"},
    {"id": "sam.okafor",  "name": "Sam Okafor",   "role": "manager", "group": "Alpha Leadership Team", "authority": 4, "active": "TRUE"},
    {"id": "morgan.kim",  "name": "Morgan Kim",   "role": "ic",      "group": "Alpha Leadership Team", "authority": 2, "active": "TRUE"},
    {"id": "casey.walsh", "name": "Casey Walsh",  "role": "advisor", "group": "Alpha Leadership Team", "authority": 3, "active": "TRUE"},
]

# Big Five: O, C, E, A, N — direct 0..100 entry per contract
OCEAN = [
    ("alex.rivera",  72, 81, 78, 58, 35, "self_report"),
    ("jordan.chen",  65, 88, 52, 71, 42, "self_report"),
    ("sam.okafor",   61, 74, 90, 68, 38, "self_report"),
    ("morgan.kim",   79, 77, 41, 65, 55, "self_report"),
    ("casey.walsh",  83, 69, 63, 76, 28, "validated"),
]

# Conflict style: Competing, Collaborating, Compromising, Avoiding, Accommodating
CONFLICT = [
    ("alex.rivera",  35, 30, 20,  5, 10, "observed"),
    ("jordan.chen",  15, 45, 25, 10,  5, "observed"),
    ("sam.okafor",   40, 25, 20,  5, 10, "observed"),
    ("morgan.kim",   10, 35, 25, 25,  5, "self_report"),
    ("casey.walsh",  10, 40, 30,  5, 15, "self_report"),
]

# Psych Safety: 7 Edmondson items, scale 1..5
PSYCH_SAFETY = [
    ("alex.rivera",  4, 4, 3, 4, 4, 3, 3, "self_report"),
    ("jordan.chen",  4, 4, 4, 3, 4, 4, 4, "self_report"),
    ("sam.okafor",   3, 3, 3, 3, 3, 3, 3, "self_report"),
    ("morgan.kim",   2, 3, 2, 2, 3, 2, 3, "self_report"),
    ("casey.walsh",  4, 5, 4, 5, 4, 4, 4, "self_report"),
]

# Comm/Decision: directness, context_orientation, verbal_dominance, listening_quality,
#                feedback_tolerance, analytical_vs_intuitive, risk_appetite,
#                decision_speed, ambiguity_tolerance
COMM = [
    ("alex.rivera",  85, 35, 78, 55, 62, 55, 68, 80, 60, "observed"),
    ("jordan.chen",  72, 42, 45, 78, 71, 85, 42, 58, 45, "observed"),
    ("sam.okafor",   90, 28, 88, 51, 58, 35, 78, 88, 72, "observed"),
    ("morgan.kim",   52, 55, 30, 82, 68, 88, 38, 42, 52, "self_report"),
    ("casey.walsh",  65, 68, 55, 88, 80, 62, 50, 62, 75, "observed"),
]

# EQ: perceiving, using, understanding, managing
EQ = [
    ("alex.rivera",  68, 70, 65, 72, "self_report"),
    ("jordan.chen",  75, 68, 78, 72, "self_report"),
    ("sam.okafor",   62, 72, 55, 58, "self_report"),
    ("morgan.kim",   78, 55, 72, 60, "self_report"),
    ("casey.walsh",  82, 78, 80, 85, "validated"),
]

# Attachment: secure, anxious, avoidant, fearful (each row sums to 100)
ATTACHMENT = [
    ("alex.rivera",  55, 20, 15, 10, "inferred"),
    ("jordan.chen",  65, 18, 12,  5, "inferred"),
    ("sam.okafor",   45, 30, 15, 10, "inferred"),
    ("morgan.kim",   40, 35, 15, 10, "inferred"),
    ("casey.walsh",  70, 15, 10,  5, "inferred"),
]

# Gate A validation checks (17 required)
VALIDATION_CHECKS = [
    ("GA-01", "People",          "All 5 PersonIDs match slug pattern (lowercase, no spaces, 3–64 chars)",   "ID_Valid=TRUE all rows"),
    ("GA-02", "People",          "No duplicate PersonIDs",                                                   "0 duplicates"),
    ("GA-03", "People",          "Role values all in allowed enum",                                          "All valid"),
    ("GA-04", "People",          "AuthorityLevel all in 1..5",                                               "All valid"),
    ("GA-05", "Big Five",        "All OCEAN values in 0..100",                                               "No red cells"),
    ("GA-06", "Big Five",        "No missing-data flags (all 5 fields present per person)",                  "Missing_Flag=FALSE"),
    ("GA-07", "Conflict Style",  "Mode sums all within 99–101",                                              "Mode_Sum in [99,101]"),
    ("GA-08", "Conflict Style",  "Sum_Valid = TRUE for all 5 rows",                                          "Sum_Valid=TRUE"),
    ("GA-09", "Psych Safety",    "All item values in 1..5",                                                  "No red cells"),
    ("GA-10", "Psych Safety",    "Per-person scores match spec ±1: 64, 71, 50, 36, 82",                     "Within ±1 of spec"),
    ("GA-11", "Psych Safety",    "Group aggregate = 61 ±1",                                                  "60–62"),
    ("GA-12", "Comm/Decision",   "All values in 0..100",                                                     "No red cells"),
    ("GA-13", "Comm/Decision",   "Polarity labels present in header row 2",                                  "Labels visible"),
    ("GA-14", "EQ",              "All values in 0..100; EQ_Composite computed for all rows",                 "Composite filled, no red"),
    ("GA-15", "Attachment",      "Attach_Sum = 100 and Sum_Valid = TRUE for all 5 rows",                     "Sum_Valid=TRUE"),
    ("GA-16", "Cross-tab",       "PersonID in each assessment tab resolves to valid People!PersonID",         "No NOT FOUND"),
    ("GA-17", "Workbook",        f'Sheet version tag "{VERSION}" present in README-Consent A1',              "Tag present"),
]

# Gate B synthetic rows
REL_METRIC_BASE = {
    "alex.rivera":  [82, 88, 70, 81, 25, 42, 85, 14, 58, 72],
    "jordan.chen":  [78, 64, 68, 76, 31, 38, 77, 21, 44, 55],
    "sam.okafor":   [65, 79, 60, 66, 48, 35, 80, 19, 51, 47],
    "morgan.kim":   [59, 45, 56, 61, 40, 50, 63, 33, 36, 28],
    "casey.walsh":  [84, 57, 75, 83, 22, 41, 74, 13, 62, 34],
}

GROUP_CONTEXT = {
    "group.id": "alpha-leadership-team",
    "group.type": "leadership_team",
    "group.structure": "formal",
    "group.shared_goals": "Deliver Q3 platform migration while preserving customer trust and team health.",
    "group.norms.explicit": "Weekly decision sync; written RFCs for material changes; blameless retrospectives.",
    "group.norms.implicit": "Defer to domain expert in meetings; avoid public disagreement with executive sponsor.",
    "group.decision_rules": "Consensus-seeking with leader tie-break on deadline-critical calls.",
    "group.conflict_history": "Two unresolved disputes on resourcing and release scope in previous quarter.",
    "group.stress_level": 4,
    "group.role_clarity": 72,
    "group.cultural_context": "Hybrid US-based team with direct communication norm and high execution urgency.",
    "group.environmental_constraints": "Fixed launch date, constrained staffing, and elevated executive visibility.",
}

SCENARIO_CONTEXT = {
    "scenario.id": "q3-migration-release-risk",
    "scenario.title": "High-risk launch scope decision",
    "scenario.type": "decision",
    "scenario.trigger_event": (
        "Security testing found late-breaking vulnerabilities in two non-critical modules "
        "48 hours before launch."
    ),
    "scenario.stakes_level": 5,
    "scenario.emotional_intensity": 4,
    "scenario.ambiguity_level": 4,
    "scenario.time_pressure": 5,
    "scenario.resource_constraints": "No additional engineering headcount can be added before launch date.",
    "scenario.public_visibility": "TRUE",
    "scenario.required_decision": "Choose between delaying launch, reducing scope, or shipping with mitigations.",
    "scenario.success_criteria": "Protect customer trust and security while meeting critical business commitments.",
    "scenario.failure_consequences": "Customer churn risk, executive escalation, and increased team burnout.",
    "scenario.known_facts": "- Two modules fail security threshold\n- Core product path passes\n- Communications plan drafted",
    "scenario.uncertain_facts": "- Exploitability timeline\n- Customer tolerance for delay\n- Vendor patch ETA",
    "scenario.intervention_options": "- Facilitated decision protocol\n- Rapid risk triage pod\n- Stakeholder expectation reset",
}

# ── Style constants ───────────────────────────────────────────────────────────

_THIN   = Side(style="thin")
_MEDIUM = Side(style="medium")

HDR_FILL      = PatternFill("solid", fgColor="2F5496")
SUBHDR_FILL   = PatternFill("solid", fgColor="BDD7EE")
COMPUTED_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL      = PatternFill("solid", fgColor="FFCCCC")
YELLOW_FILL   = PatternFill("solid", fgColor="FFFFC7")
GRAY_FILL     = PatternFill("solid", fgColor="F2F2F2")

HDR_FONT    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT   = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
SMALL_FONT  = Font(name="Calibri", size=9, italic=True, color="595959")
NOTE_FONT   = Font(name="Calibri", size=9, color="595959")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# ── Style helpers ─────────────────────────────────────────────────────────────

def _h(ws, row, col, text, width=None):
    """Write a header cell."""
    c = ws.cell(row=row, column=col, value=text)
    c.font, c.fill, c.alignment, c.border = HDR_FONT, HDR_FILL, CENTER, CELL_BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _sh(ws, row, col, text, width=None):
    """Write a sub-header / polarity-label cell."""
    c = ws.cell(row=row, column=col, value=text)
    c.font, c.fill, c.alignment, c.border = NOTE_FONT, SUBHDR_FILL, CENTER, CELL_BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _d(ws, row, col, value=None):
    """Write a data cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.border, c.alignment = NORMAL_FONT, CELL_BORDER, LEFT
    return c

def _f(ws, row, col, formula):
    """Write a computed (formula) cell."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font, c.fill, c.border, c.alignment = NORMAL_FONT, COMPUTED_FILL, CELL_BORDER, CENTER
    return c

def _add_dv_list(ws, cell_range, choices, msg=""):
    dv = DataValidation(
        type="list",
        formula1=f'"{choices}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error=msg or f"Must be one of: {choices}",
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)

def _add_dv_int(ws, cell_range, lo, hi):
    dv = DataValidation(
        type="whole", operator="between",
        formula1=str(lo), formula2=str(hi),
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Out of range",
        error=f"Enter an integer between {lo} and {hi}.",
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)

def _red_if(ws, cell_range, formula):
    """Apply red fill when the formula is TRUE."""
    ws.conditional_formatting.add(cell_range, FormulaRule(formula=[formula], fill=RED_FILL))

def _yellow_if(ws, cell_range, formula):
    ws.conditional_formatting.add(cell_range, FormulaRule(formula=[formula], fill=YELLOW_FILL))

def _col(n):
    return get_column_letter(n)

# ── Tab builders ──────────────────────────────────────────────────────────────

def build_readme(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80

    def section(row, label, text):
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=row, column=2, value=text)
        c.font, c.alignment = NORMAL_FONT, LEFT
        ws.row_dimensions[row].height = 30

    # Version tag — GA-17 checks this cell
    tag = ws.cell(row=1, column=1, value=f"version: {VERSION}")
    tag.font = SMALL_FONT
    ws.cell(row=1, column=2, value=f"created: {TODAY}").font = SMALL_FONT

    ws.cell(row=2, column=1).fill = GRAY_FILL

    section(3,  "Project",
            "Group Dynamics Simulator — Phase 1 Spreadsheet Prototype. "
            "Personal use only. Validates the psychological model and prompt design before Phase 2 software build.")
    section(5,  "Purpose",
            "Given structured psychological profiles, relationship data, and a scenario, simulate probable "
            "group dynamics and produce probability-weighted, evidence-traced outcomes.")
    section(7,  "NOT for",
            "Clinical diagnosis · Hiring or performance decisions · Legal proceedings · Multi-tenant use.")
    section(9,  "Consent",
            "All assessment data should be collected with the knowledge of the individuals assessed, "
            "or clearly marked as observer estimates / inferences.")
    section(11, "Evidence sources",
            "validated — instrument-administered and quality-checked  |  "
            "self_report — self-entered by subject  |  "
            "observed — coach/facilitator estimate from direct observation  |  "
            "inferred — reasoned from indirect evidence  |  "
            "missing — not collected")
    section(13, "Tab guide",
            "People → Big Five → Conflict Style → Psych Safety → Comm-Decision → EQ → Attachment  "
            "(assessment inputs) · Relationship Matrix → Group Context → Scenario Builder → Simulation Config  "
            "(Gate B inputs)  ·  Gate-A-Validation-Log (quality checks)")
    section(15, "Version log",
            "Edit the table below each session.")

    # Version log header
    for col, text in enumerate(["Date", "Editor", "Change description"], 1):
        _h(ws, 17, col, text)
    _d(ws, 18, 1, TODAY)
    _d(ws, 18, 2, "build_workbook.py")
    _d(ws, 18, 3, f"Workbook generated; version {VERSION}; Gate A + Gate B synthetic dataset loaded.")


def build_people(ws):
    headers = [
        ("PersonID",         16),
        ("DisplayName",      20),
        ("Role",             13),
        ("GroupMembership",  24),
        ("AuthorityLevel",   14),
        ("IsActive",         10),
        ("ID_Valid",         10),
    ]
    for col, (text, width) in enumerate(headers, 1):
        _h(ws, 1, col, text, width)

    DATA_ROWS = range(2, 7)
    for i, p in enumerate(PEOPLE):
        r = i + 2
        _d(ws, r, 1, p["id"])
        _d(ws, r, 2, p["name"])
        _d(ws, r, 3, p["role"])
        _d(ws, r, 4, p["group"])
        _d(ws, r, 5, p["authority"])
        _d(ws, r, 6, p["active"])
        # Simplified slug check: lowercase, no spaces, length 3–64
        _f(ws, r, 7,
           f'=AND(LEN(A{r})>=3,LEN(A{r})<=64,EXACT(LOWER(A{r}),A{r}),ISERROR(FIND(" ",A{r})))')

    _add_dv_list(ws, "C2:C6", "leader,manager,ic,advisor,observer,other")
    _add_dv_list(ws, "F2:F6", "TRUE,FALSE")
    _add_dv_int(ws,  "E2:E6", 1, 5)

    # Red if ID_Valid is FALSE
    _red_if(ws, "A2:A6", "NOT(G2)")
    # Red if AuthorityLevel out of range
    _red_if(ws, "E2:E6", "OR(E2<1,E2>5)")


def build_big_five(ws):
    traits = ["Openness", "Conscientiousness", "Extraversion", "Agreeableness", "Neuroticism"]
    hdrs = [("PersonID", 14)] + [(t, 16) for t in traits] + [("Evidence Source", 16), ("Missing Flag", 12)]
    for col, (text, width) in enumerate(hdrs, 1):
        _h(ws, 1, col, text, width)

    # Polarity note row 2
    _sh(ws, 2, 1, "cross-ref → People!A")
    for col in range(2, 7):
        _sh(ws, 2, col, "direct 0–100 entry")
    _sh(ws, 2, 7, "source enum")
    _sh(ws, 2, 8, "computed")

    for i, row_data in enumerate(OCEAN):
        r = i + 3
        pid, O, C, E, A, N, src = row_data
        _d(ws, r, 1, pid)
        for col, val in enumerate([O, C, E, A, N], 2):
            _d(ws, r, col, val)
        _d(ws, r, 7, src)
        _f(ws, r, 8, f'=IF(A{r}="","",COUNTA(B{r}:F{r})<5)')

    _add_dv_int(ws, "B3:F7", 0, 100)
    _add_dv_list(ws, "G3:G7", "validated,self_report,observed,inferred,missing")
    # Red on out-of-range trait values
    _red_if(ws, "B3:F7", "OR(B3<0,B3>100)")
    # Yellow on missing flag = TRUE
    _yellow_if(ws, "H3:H7", "H3=TRUE")


def build_conflict(ws):
    modes = ["Competing", "Collaborating", "Compromising", "Avoiding", "Accommodating"]
    hdrs = [("PersonID", 14)] + [(m, 14) for m in modes] + \
           [("Mode Sum", 10), ("Sum Valid\n(100±1)", 12), ("Evidence Source", 16)]
    for col, (text, width) in enumerate(hdrs, 1):
        _h(ws, 1, col, text, width)

    for i, row_data in enumerate(CONFLICT):
        r = i + 2
        pid, comp, collab, compromise, avoid, accom, src = row_data
        _d(ws, r, 1, pid)
        for col, val in enumerate([comp, collab, compromise, avoid, accom], 2):
            _d(ws, r, col, val)
        _f(ws, r, 7, f"=SUM(B{r}:F{r})")
        _f(ws, r, 8, f'=IF(G{r}=0,"",ABS(G{r}-100)<=1)')
        _d(ws, r, 9, src)

    _add_dv_int(ws, "B2:F6", 0, 100)
    _add_dv_list(ws, "I2:I6", "validated,self_report,observed,inferred,missing")
    # Red row when sum is outside 99–101
    _red_if(ws, "A2:I6", 'AND($G2<>"",ABS($G2-100)>1)')


def build_psych_safety(ws):
    hdrs = [("PersonID", 14)] + [(f"Item {i}", 8) for i in range(1, 8)] + \
           [("PS Score\n(0–100)", 12), ("Missing\nFlag", 10), ("Evidence Source", 16)]
    for col, (text, width) in enumerate(hdrs, 1):
        _h(ws, 1, col, text, width)

    _sh(ws, 2, 1, "cross-ref → People!A")
    for col in range(2, 9):
        _sh(ws, 2, col, "1=low safety · 5=high safety")
    _sh(ws, 2, 9, "=ROUND(100×(avg−1)/4,0)")
    _sh(ws, 2, 10, "computed")
    _sh(ws, 2, 11, "source enum")

    for i, row_data in enumerate(PSYCH_SAFETY):
        r = i + 3
        pid = row_data[0]
        items = row_data[1:8]
        src = row_data[8]
        _d(ws, r, 1, pid)
        for col, val in enumerate(items, 2):
            _d(ws, r, col, val)
        _f(ws, r, 9, f'=IF(A{r}="","",ROUND(100*(AVERAGE(B{r}:H{r})-1)/4,0))')
        _f(ws, r, 10, f'=IF(A{r}="","",COUNTA(B{r}:H{r})<7)')
        _d(ws, r, 11, src)

    # Group aggregate row
    agg_row = 9
    ws.cell(row=agg_row, column=1, value="GROUP AGGREGATE").font = BOLD_FONT
    _f(ws, agg_row, 9, f"=IFERROR(ROUND(AVERAGE(I3:I7),0),\"\")")

    _add_dv_int(ws, "B3:H7", 1, 5)
    _add_dv_list(ws, "K3:K7", "validated,self_report,observed,inferred,missing")
    _red_if(ws, "B3:H7", "OR(B3<1,B3>5)")
    _yellow_if(ws, "J3:J7", "J3=TRUE")


def build_comm_decision(ws):
    comm_hdrs = [
        ("Directness\n(0=indirect,\n100=direct)", 16),
        ("Context Orient.\n(0=low-ctx,\n100=high-ctx)", 16),
        ("Verbal\nDominance", 13),
        ("Listening\nQuality", 13),
        ("Feedback\nTolerance", 13),
    ]
    dec_hdrs = [
        ("Analytical vs\nIntuitive\n(0=intuitive,\n100=analytical)", 18),
        ("Risk\nAppetite", 12),
        ("Decision\nSpeed", 12),
        ("Ambiguity\nTolerance", 13),
    ]
    all_hdrs = [("PersonID", 14)] + comm_hdrs + dec_hdrs + \
               [("Evidence Source", 16), ("Missing\nFlag", 10)]

    for col, (text, width) in enumerate(all_hdrs, 1):
        _h(ws, 1, col, text, width)

    # Section labels row 2
    _sh(ws, 2, 1, "cross-ref → People!A")
    for col in range(2, 7):
        _sh(ws, 2, col, "Communication")
    for col in range(7, 11):
        _sh(ws, 2, col, "Decision")
    _sh(ws, 2, 11, "source enum")
    _sh(ws, 2, 12, "computed")

    for i, row_data in enumerate(COMM):
        r = i + 3
        pid = row_data[0]
        vals = row_data[1:10]
        src = row_data[10]
        _d(ws, r, 1, pid)
        for col, val in enumerate(vals, 2):
            _d(ws, r, col, val)
        _d(ws, r, 11, src)
        # Missing if ≥5 of 9 fields blank
        _f(ws, r, 12, f'=IF(A{r}="","",COUNTA(B{r}:J{r})<5)')

    _add_dv_int(ws, "B3:J7", 0, 100)
    _add_dv_list(ws, "K3:K7", "validated,self_report,observed,inferred,missing")
    _red_if(ws, "B3:J7", "OR(B3<0,B3>100)")
    _yellow_if(ws, "L3:L7", "L3=TRUE")


def build_eq(ws):
    dims = ["Perceiving", "Using", "Understanding", "Managing"]
    hdrs = [("PersonID", 14)] + [(d, 15) for d in dims] + \
           [("EQ Composite\n(mean)", 14), ("Missing\nFlag", 10), ("Evidence Source", 16)]
    for col, (text, width) in enumerate(hdrs, 1):
        _h(ws, 1, col, text, width)

    for i, row_data in enumerate(EQ):
        r = i + 2
        pid, perc, using, under, manag, src = row_data
        _d(ws, r, 1, pid)
        for col, val in enumerate([perc, using, under, manag], 2):
            _d(ws, r, col, val)
        _f(ws, r, 6, f'=IF(A{r}="","",IFERROR(ROUND(AVERAGE(B{r}:E{r}),0),""))')
        _f(ws, r, 7, f'=IF(A{r}="","",COUNTA(B{r}:E{r})<4)')
        _d(ws, r, 8, src)

    _add_dv_int(ws, "B2:E6", 0, 100)
    _add_dv_list(ws, "H2:H6", "validated,self_report,observed,inferred,missing")
    _red_if(ws, "B2:E6", "OR(B2<0,B2>100)")
    _yellow_if(ws, "G2:G6", "G2=TRUE")


def build_attachment(ws):
    hdrs = [
        ("PersonID",        14),
        ("Secure",          12),
        ("Anxious",         12),
        ("Avoidant",        12),
        ("Fearful",         12),
        ("Attach Sum",      12),
        ("Sum Valid\n(100±1)",  14),
        ("Evidence Source", 16),
    ]
    for col, (text, width) in enumerate(hdrs, 1):
        _h(ws, 1, col, text, width)

    note = ws.cell(row=2, column=1,
                   value="Optional. Leave all four blank if not used. When all four are present they should sum to 100 ±1.")
    note.font = SMALL_FONT
    ws.merge_cells("A2:H2")
    note.alignment = LEFT

    for i, row_data in enumerate(ATTACHMENT):
        r = i + 3
        pid, sec, anx, avoid, fear, src = row_data
        _d(ws, r, 1, pid)
        for col, val in enumerate([sec, anx, avoid, fear], 2):
            _d(ws, r, col, val)
        _f(ws, r, 6, f'=IF(COUNTA(B{r}:E{r})=0,"",SUM(B{r}:E{r}))')
        _f(ws, r, 7, f'=IF(F{r}="","",IF(COUNTA(B{r}:E{r})<4,"incomplete",IF(ABS(F{r}-100)<=1,TRUE,FALSE)))')
        _d(ws, r, 8, src)

    _add_dv_int(ws, "B3:E7", 0, 100)
    _add_dv_list(ws, "H3:H7", "validated,self_report,observed,inferred,missing")
    # Red when sum is outside 99–101 and all four values are present
    _red_if(ws, "A3:H7", 'AND(COUNTA($B3:$E3)=4,ABS($F3-100)>1)')
    # Yellow when incomplete
    _yellow_if(ws, "G3:G7", 'G3="incomplete"')


def build_validation_log(ws):
    hdrs = [
        ("Check ID",         10),
        ("Tab",              18),
        ("Check Description",50),
        ("Expected",         28),
        ("Actual",           18),
        ("Pass / Fail",      12),
        ("Notes",            30),
        ("Date",             12),
        ("Reviewer",         16),
    ]
    for col, (text, width) in enumerate(hdrs, 1):
        _h(ws, 1, col, text, width)

    for i, check in enumerate(VALIDATION_CHECKS):
        r = i + 2
        check_id, tab, desc, expected = check
        _d(ws, r, 1, check_id)
        _d(ws, r, 2, tab)
        _d(ws, r, 3, desc)
        _d(ws, r, 4, expected)
        _d(ws, r, 5, "")   # Actual — fill manually
        _d(ws, r, 6, "")   # Pass/Fail — fill manually
        _d(ws, r, 7, "")   # Notes
        _d(ws, r, 8, "")   # Date
        _d(ws, r, 9, "")   # Reviewer

    _add_dv_list(ws, "F2:F18", "Pass,Fail,N/A")
    # Green fill when Pass, red when Fail
    ws.conditional_formatting.add("F2:F18",
        FormulaRule(formula=['F2="Pass"'], fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add("F2:F18",
        FormulaRule(formula=['F2="Fail"'], fill=RED_FILL))

    ws.row_dimensions[1].height = 28


def build_relationship_matrix(ws):
    hdrs = [
        ("FromPersonID", 16),
        ("ToPersonID", 16),
        ("Trust", 10),
        ("Influence", 10),
        ("Emotional Closeness", 18),
        ("Respect", 10),
        ("Conflict Intensity", 16),
        ("Dependency", 12),
        ("Comm Frequency", 14),
        ("Avoidance", 10),
        ("Alliance", 10),
        ("Power Differential", 16),
        ("Health Score", 12),
        ("Evidence Source", 16),
        ("Notes", 36),
    ]
    for col, (text, width) in enumerate(hdrs, 1):
        _h(ws, 1, col, text, width)

    row = 2
    ids = [p["id"] for p in PEOPLE]
    for from_pid in ids:
        for to_pid in ids:
            if from_pid == to_pid:
                continue
            _d(ws, row, 1, from_pid)
            _d(ws, row, 2, to_pid)
            base = REL_METRIC_BASE[from_pid]
            adjust = (ids.index(to_pid) - ids.index(from_pid)) * 2
            values = [max(0, min(100, v + adjust)) for v in base]
            for col, val in enumerate(values, 3):
                _d(ws, row, col, val)
            _f(ws, row, 13, f"=ROUND((C{row}+D{row}+E{row}+F{row}+I{row}+K{row}-(G{row}+J{row}+L{row}))/6,0)")
            _d(ws, row, 14, "observed")
            _d(ws, row, 15, "Synthetic Gate B edge seed.")
            row += 1

    last_row = row - 1
    ws.cell(row=last_row + 2, column=1, value="NETWORK HEALTH").font = BOLD_FONT
    _f(ws, last_row + 2, 13, f'=IFERROR(ROUND(AVERAGE(M2:M{last_row}),0),"")')

    _add_dv_int(ws, f"C2:L{last_row}", 0, 100)
    _add_dv_list(ws, f"N2:N{last_row}", "validated,self_report,observed,inferred,missing")
    _red_if(ws, f"A2:O{last_row}", "$A2=$B2")
    _red_if(ws, f"C2:L{last_row}", "OR(C2<0,C2>100)")


def build_group_context(ws):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 95
    ws.column_dimensions["C"].width = 28
    _h(ws, 1, 1, "Field")
    _h(ws, 1, 2, "Value")
    _h(ws, 1, 3, "Rule")

    rows = [
        ("group.id", GROUP_CONTEXT["group.id"], "required; slug"),
        ("group.type", GROUP_CONTEXT["group.type"], "team,leadership_team,project_team,board,other"),
        ("group.structure", GROUP_CONTEXT["group.structure"], "formal,informal,hybrid"),
        ("group.shared_goals", GROUP_CONTEXT["group.shared_goals"], "required"),
        ("group.norms.explicit", GROUP_CONTEXT["group.norms.explicit"], "optional"),
        ("group.norms.implicit", GROUP_CONTEXT["group.norms.implicit"], "optional"),
        ("group.psychological_safety_aggregate", '=IFERROR(ROUND(AVERAGE(\'Psych Safety\'!I3:I7),0),"")', "computed read-only"),
        ("group.decision_rules", GROUP_CONTEXT["group.decision_rules"], "optional"),
        ("group.conflict_history", GROUP_CONTEXT["group.conflict_history"], "optional"),
        ("group.stress_level", GROUP_CONTEXT["group.stress_level"], "required 1..5"),
        ("group.role_clarity", GROUP_CONTEXT["group.role_clarity"], "optional 0..100"),
        ("group.cultural_context", GROUP_CONTEXT["group.cultural_context"], "optional"),
        ("group.environmental_constraints", GROUP_CONTEXT["group.environmental_constraints"], "optional"),
    ]

    for r, (field, value, rule) in enumerate(rows, start=2):
        _d(ws, r, 1, field)
        if isinstance(value, str) and value.startswith("="):
            _f(ws, r, 2, value)
        else:
            _d(ws, r, 2, value)
        _d(ws, r, 3, rule)

    _add_dv_list(ws, "B3:B3", "team,leadership_team,project_team,board,other")
    _add_dv_list(ws, "B4:B4", "formal,informal,hybrid")
    _add_dv_int(ws, "B11:B11", 1, 5)
    _add_dv_int(ws, "B12:B12", 0, 100)


def build_scenario_builder(ws):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    ws.column_dimensions["C"].width = 30
    _h(ws, 1, 1, "Field")
    _h(ws, 1, 2, "Value")
    _h(ws, 1, 3, "Rule")

    rows = [
        ("scenario.id", SCENARIO_CONTEXT["scenario.id"], "required; slug"),
        ("scenario.title", SCENARIO_CONTEXT["scenario.title"], "required"),
        ("scenario.type", SCENARIO_CONTEXT["scenario.type"], "conflict,decision,change,crisis,performance,other"),
        ("scenario.trigger_event", SCENARIO_CONTEXT["scenario.trigger_event"], "required"),
        ("scenario.stakes_level", SCENARIO_CONTEXT["scenario.stakes_level"], "required 1..5"),
        ("scenario.emotional_intensity", SCENARIO_CONTEXT["scenario.emotional_intensity"], "required 1..5"),
        ("scenario.ambiguity_level", SCENARIO_CONTEXT["scenario.ambiguity_level"], "required 1..5"),
        ("scenario.time_pressure", SCENARIO_CONTEXT["scenario.time_pressure"], "required 1..5"),
        ("scenario.resource_constraints", SCENARIO_CONTEXT["scenario.resource_constraints"], "optional"),
        ("scenario.public_visibility", SCENARIO_CONTEXT["scenario.public_visibility"], "TRUE,FALSE"),
        ("scenario.required_decision", SCENARIO_CONTEXT["scenario.required_decision"], "required"),
        ("scenario.success_criteria", SCENARIO_CONTEXT["scenario.success_criteria"], "required"),
        ("scenario.failure_consequences", SCENARIO_CONTEXT["scenario.failure_consequences"], "required"),
        ("scenario.known_facts", SCENARIO_CONTEXT["scenario.known_facts"], "optional newline list"),
        ("scenario.uncertain_facts", SCENARIO_CONTEXT["scenario.uncertain_facts"], "optional newline list"),
        ("scenario.intervention_options", SCENARIO_CONTEXT["scenario.intervention_options"], "optional newline list"),
    ]

    for r, (field, value, rule) in enumerate(rows, start=2):
        _d(ws, r, 1, field)
        _d(ws, r, 2, value)
        _d(ws, r, 3, rule)

    _add_dv_list(ws, "B4:B4", "conflict,decision,change,crisis,performance,other")
    _add_dv_int(ws, "B6:B9", 1, 5)
    _add_dv_list(ws, "B11:B11", "TRUE,FALSE")


def build_simulation_config(ws):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 36
    _h(ws, 1, 1, "Field")
    _h(ws, 1, 2, "Value")
    _h(ws, 1, 3, "Rule")

    run_id_formula = '=LOWER(TEXT(NOW(),"yyyymmdd-hhmmss")&"-gds")'
    rows = [
        ("sim.run_id", run_id_formula, "required, unique (timestamp suffix)"),
        ("sim.prompt_version_key", "P1.0", "required regex P<major>.<minor>"),
        ("sim.passes", 3, "1..10"),
        ("sim.randomness", "medium", "low,medium,high"),
        ("sim.depth", "standard", "surface,standard,deep"),
        ("sim.dialogue_enabled", "TRUE", "TRUE,FALSE"),
        ("sim.report_detail_level", "standard", "summary,standard,full"),
        ("sim.intervention_mode", "baseline", "baseline,compare"),
        ("sim.evidence_strictness", "strict", "strict,moderate,lenient"),
        ("sim.guardrail_verbosity", "standard", "minimal,standard,verbose"),
    ]

    for r, (field, value, rule) in enumerate(rows, start=2):
        _d(ws, r, 1, field)
        if isinstance(value, str) and value.startswith("="):
            _f(ws, r, 2, value)
        else:
            _d(ws, r, 2, value)
        _d(ws, r, 3, rule)

    _add_dv_int(ws, "B4:B4", 1, 10)
    _add_dv_list(ws, "B5:B5", "low,medium,high")
    _add_dv_list(ws, "B6:B6", "surface,standard,deep")
    _add_dv_list(ws, "B7:B7", "TRUE,FALSE")
    _add_dv_list(ws, "B8:B8", "summary,standard,full")
    _add_dv_list(ws, "B9:B9", "baseline,compare")
    _add_dv_list(ws, "B10:B10", "strict,moderate,lenient")
    _add_dv_list(ws, "B11:B11", "minimal,standard,verbose")


def build_structured_profile_output(ws):
    headers = [
        ("PersonID", 14),
        ("DisplayName", 18),
        ("Role", 12),
        ("OCEAN Summary", 42),
        ("Dominant Conflict Mode", 20),
        ("Communication Pattern", 28),
        ("Decision Pattern", 28),
        ("EQ Summary", 20),
        ("Attachment Tendency", 20),
        ("Confidence", 12),
        ("Missing Data", 12),
    ]
    for col, (text, width) in enumerate(headers, 1):
        _h(ws, 1, col, text, width)

    ws.row_dimensions[1].height = 24

    for r in range(2, 7):
        _d(ws, r, 1, f'=People!A{r}')
        _d(ws, r, 2, f'=People!B{r}')
        _d(ws, r, 3, f'=People!C{r}')
        _d(
            ws, r, 4,
            (
                f'="O:"&IF(\'Big Five\'!B{r+1}>=67,"high",IF(\'Big Five\'!B{r+1}>=34,"medium","low"))'
                f'&", C:"&IF(\'Big Five\'!C{r+1}>=67,"high",IF(\'Big Five\'!C{r+1}>=34,"medium","low"))'
                f'&", E:"&IF(\'Big Five\'!D{r+1}>=67,"high",IF(\'Big Five\'!D{r+1}>=34,"medium","low"))'
                f'&", A:"&IF(\'Big Five\'!E{r+1}>=67,"high",IF(\'Big Five\'!E{r+1}>=34,"medium","low"))'
                f'&", N:"&IF(\'Big Five\'!F{r+1}>=67,"high",IF(\'Big Five\'!F{r+1}>=34,"medium","low"))'
            )
        )
        _d(
            ws, r, 5,
            (
                f'=INDEX(\'Conflict Style\'!$B$1:$F$1,'
                f'MATCH(MAX(\'Conflict Style\'!B{r}:F{r}),\'Conflict Style\'!B{r}:F{r},0))'
            )
        )
        _d(
            ws, r, 6,
            (
                f'="Directness "&\'Comm-Decision\'!B{r+1}&"/100; Listening "&\'Comm-Decision\'!E{r+1}'
                f'&"/100; Feedback tol "&\'Comm-Decision\'!F{r+1}&"/100"'
            )
        )
        _d(
            ws, r, 7,
            (
                f'="Analytical "&\'Comm-Decision\'!G{r+1}&"/100; Risk "&\'Comm-Decision\'!H{r+1}'
                f'&"/100; Speed "&\'Comm-Decision\'!I{r+1}&"/100"'
            )
        )
        _d(
            ws, r, 8,
            (
                f'="Composite "&EQ!F{r}&"/100 ("'
                f'&IF(EQ!F{r}>=67,"high",IF(EQ!F{r}>=34,"medium","low"))&")"'
            )
        )
        _d(
            ws, r, 9,
            (
                f'=IF(Attachment!G{r+1}<>"TRUE","check attachment quality",'
                f'INDEX(Attachment!$B$1:$E$1,MATCH(MAX(Attachment!B{r+1}:E{r+1}),Attachment!B{r+1}:E{r+1},0)))'
            )
        )
        _d(ws, r, 10, f'=IF(OR(\'Big Five\'!H{r+1},\'Comm-Decision\'!L{r+1},EQ!G{r}),"moderate","high")')
        _d(ws, r, 11, f'=OR(\'Big Five\'!H{r+1},\'Psych Safety\'!J{r+1},\'Comm-Decision\'!L{r+1},EQ!G{r})')

    _yellow_if(ws, "K2:K6", "K2=TRUE")


def build_prompt_inputs(ws):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 140
    _h(ws, 1, 1, "Section")
    _h(ws, 1, 2, "Prompt Text")

    sections = [
        ("SYSTEM ROLE", (
            "You are an organizational group dynamics simulator for coaching/research reflection (not diagnosis, "
            "hiring, legal, or disciplinary use). Maintain explicit separation of EVIDENCE, INFERENCE, and "
            "SIMULATION layers. Use cautious probabilistic language and uncertainty statements. Reject deterministic "
            "claims and avoid pathologizing labels."
        )),
        ("PERSON PROFILES", (
            "For each person_id from Structured Profile Output, create a compact block: role, OCEAN low/medium/high "
            "summary, dominant conflict mode, communication and decision pattern, EQ composite summary, attachment "
            "tendency (or missing), confidence category, and missing-data flags. For each non-trivial claim, reference "
            "which input column(s) support it."
        )),
        ("RELATIONSHIP DATA", (
            "Use directed pair edges from Relationship Matrix: trust, influence, emotional_closeness, respect, "
            "conflict_intensity, dependency, communication_frequency, avoidance, alliance, power_differential, "
            "plus computed health score. Highlight asymmetries (A->B differs from B->A), high-conflict dyads, "
            "and likely coalition fault lines."
        )),
        ("GROUP CONTEXT", (
            "Use Group Context fields exactly as listed (type, structure, goals, norms, decision rules, conflict "
            "history, stress level, role clarity, cultural context, constraints). Treat these as system-level "
            "boundary conditions for behavior interpretation."
        )),
        ("SCENARIO", (
            "Use Scenario Builder exactly as listed. Restate trigger event, stakes, emotional intensity, ambiguity, "
            "time pressure, required decision, success criteria, and failure consequences before simulation."
        )),
        ("SIMULATION CONFIG", (
            "Use Simulation Config exactly as listed (passes, randomness, depth, dialogue_enabled, report detail, "
            "intervention mode, evidence strictness, guardrail verbosity). Honor these controls literally."
        )),
        ("SIMULATION PROCEDURE", (
            "Run N passes. For each pass output steps: (1) private appraisal by each actor, (2) first public move, "
            "(3) interaction sequence and influence shifts, (4) conflict escalation/de-escalation markers, "
            "(5) decision or non-decision point, (6) short-term outcome classification, "
            "(7) evidence-consistency check with uncertainties."
        )),
        ("OUTPUT REQUIREMENTS", (
            "Return markdown sections: Executive Summary; Scenario and Data Inputs; Behavior Trace by pass; "
            "Outcome Clusters with probabilities; Individual Behavior Forecasts; Intervention Options "
            "(baseline vs compare if enabled); Limitations & Ethics. Must include: confidence statement, "
            "limitations statement, non-determinism disclaimer, and explicit evidence-vs-inference-vs-simulation "
            "separation check."
        )),
        ("OUTPUT FORMAT CONTRACT", (
            "Also emit a JSON object with keys: run_id, prompt_version_key, scenario_title, outcome_clusters[], "
            "individual_forecasts[], intervention_recommendations[], confidence_statement, limitation_statement, "
            "layer_separation_check, and rubric_self_check."
        )),
        ("EVALUATOR RUBRIC PROMPT", (
            "After simulation output, score 1-5 with short rationale for: evidence anchoring, internal consistency, "
            "plausibility, intervention usefulness, and uncertainty quality. Compute rubric_average_score (2 decimals). "
            "Scoring anchors: 1=poor/missing, 3=adequate/mixed, 5=strong/explicit and traceable."
        )),
    ]

    for idx, (name, text) in enumerate(sections, start=2):
        _d(ws, idx, 1, name)
        _d(ws, idx, 2, text)

    prompt_row = len(sections) + 3
    _h(ws, prompt_row, 1, "Copy/Paste Prompt")
    _f(
        ws, prompt_row, 2,
        (
            '=TEXTJOIN(CHAR(10)&CHAR(10),TRUE,'
            '"SYSTEM ROLE:"&B2,'
            '"PERSON PROFILES:"&B3,'
            '"RELATIONSHIP DATA:"&B4,'
            '"GROUP CONTEXT:"&B5,'
            '"SCENARIO:"&B6,'
            '"SIMULATION CONFIG:"&B7,'
            '"SIMULATION PROCEDURE:"&B8,'
            '"OUTPUT REQUIREMENTS:"&B9,'
            '"OUTPUT FORMAT CONTRACT:"&B10,'
            '"EVALUATOR RUBRIC PROMPT:"&B11,'
            '"PROMPT_VERSION_KEY: "&\'Simulation Config\'!B3,'
            '"RUN_ID: "&\'Simulation Config\'!B2)'
        )
    )
    ws.row_dimensions[prompt_row].height = 220
    ws[f"B{prompt_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def build_simulation_output_log(ws):
    headers = [
        ("RunID", 18),
        ("Date", 12),
        ("ScenarioTitle", 24),
        ("PromptVersionKey", 14),
        ("PassNumber", 10),
        ("RawOutput", 34),
        ("OutcomeClassification", 18),
        ("ProbabilityEstimate", 14),
        ("ConfidenceEstimate", 14),
        ("KeyFindings", 26),
        ("Recommendations", 26),
        ("LimitationNotes", 24),
        ("ReviewedBy", 16),
        ("EvidenceAnchoringScore", 14),
        ("InternalConsistencyScore", 14),
        ("PlausibilityScore", 14),
        ("InterventionUsefulnessScore", 16),
        ("UncertaintyQualityScore", 15),
        ("RubricAverageScore", 14),
        ("RubricNotes", 24),
    ]
    for col, (text, width) in enumerate(headers, 1):
        _h(ws, 1, col, text, width)

    for r in range(2, 22):
        _f(ws, r, 1, '=IF($E{0}="","",\'Simulation Config\'!$B$2)'.format(r))
        _f(ws, r, 2, '=IF($E{0}="","",TODAY())'.format(r))
        _f(ws, r, 3, '=IF($E{0}="","",\'Scenario Builder\'!$B$3)'.format(r))
        _f(ws, r, 4, '=IF($E{0}="","",\'Simulation Config\'!$B$3)'.format(r))
        _d(ws, r, 5, "")
        for c in range(6, 14):
            _d(ws, r, c, "")
        for c in range(14, 19):
            _d(ws, r, c, "")
        _f(ws, r, 19, f'=IF(COUNTA(N{r}:R{r})=5,ROUND(AVERAGE(N{r}:R{r}),2),"")')
        _d(ws, r, 20, "")

    _add_dv_int(ws, "E2:E21", 1, 10)
    _add_dv_int(ws, "N2:R21", 1, 5)
    _red_if(ws, "N2:R21", "OR(N2<1,N2>5)")


def build_visuals(ws):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 88
    _h(ws, 1, 1, "Visual")
    _h(ws, 1, 2, "Status / Notes")
    rows = [
        ("Relationship trust heat map", "Use Relationship Matrix Trust column with conditional color scale."),
        ("Influence sorted bar chart", "Create from Relationship Matrix Influence by from/to aggregation."),
        ("OCEAN radar (per person)", "Use Big Five rows as chart series."),
        ("Conflict style stacked bar", "Use Conflict Style mode columns."),
        ("Outcome distribution chart", "Use Simulation Output Log OutcomeClassification counts."),
    ]
    for idx, (title, note) in enumerate(rows, start=2):
        _d(ws, idx, 1, title)
        _d(ws, idx, 2, note)


# ── Workbook assembly ─────────────────────────────────────────────────────────

TAB_ORDER = [
    ("README-Consent",        build_readme),
    ("People",                build_people),
    ("Big Five",              build_big_five),
    ("Conflict Style",        build_conflict),
    ("Psych Safety",          build_psych_safety),
    ("Comm-Decision",         build_comm_decision),
    ("EQ",                    build_eq),
    ("Attachment",            build_attachment),
    ("Relationship Matrix",   build_relationship_matrix),
    ("Group Context",         build_group_context),
    ("Scenario Builder",      build_scenario_builder),
    ("Simulation Config",     build_simulation_config),
    ("Structured Profile Output", build_structured_profile_output),
    ("Prompt Inputs",         build_prompt_inputs),
    ("Simulation Output Log", build_simulation_output_log),
    ("Visuals",               build_visuals),
    ("Gate-A-Validation-Log", build_validation_log),
]

TAB_COLORS = {
    "README-Consent":        "808080",
    "People":                "2F5496",
    "Big Five":              "375623",
    "Conflict Style":        "843C0C",
    "Psych Safety":          "1F4E79",
    "Comm-Decision":         "7030A0",
    "EQ":                    "4BACC6",
    "Attachment":            "9C6500",
    "Relationship Matrix":   "C00000",
    "Group Context":         "1D6F42",
    "Scenario Builder":      "7F6000",
    "Simulation Config":     "5B9BD5",
    "Structured Profile Output": "2E75B6",
    "Prompt Inputs":         "548235",
    "Simulation Output Log": "9E480E",
    "Visuals":               "8064A2",
    "Gate-A-Validation-Log": "404040",
}


def build():
    wb = Workbook()
    wb.remove(wb.active)          # drop the default empty sheet

    for name, builder in TAB_ORDER:
        ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = TAB_COLORS.get(name, "000000")
        ws.freeze_panes = "B2"
        builder(ws)

    # Named range: PersonIDs → People!A2:A6
    dn = DefinedName("PersonIDs", attr_text="'People'!$A$2:$A$6")
    wb.defined_names["PersonIDs"] = dn

    out_dir = os.path.join(os.path.dirname(__file__), "..", "workbook")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "group-dynamics-simulator-phase1.xlsx")
    wb.save(out_path)
    print(f"Saved: {os.path.abspath(out_path)}")


if __name__ == "__main__":
    build()
